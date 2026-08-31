#!/usr/bin/env python3
import serial
import time
from openpyxl import Workbook
import os
import sys
from datetime import datetime

# ============================================================================
# CONFIGURATION
# ============================================================================
ARDUINO_PORT = '/dev/ttyACM0'
BAUD_RATE = 9600
LOG_FILE = 'access_log.xlsx'

# ============================================================================
# COLORS
# ============================================================================
GREEN = '\033[92m'
RED = '\033[91m'
YELLOW = '\033[93m'
BLUE = '\033[94m'
CYAN = '\033[96m'
BOLD = '\033[1m'
RESET = '\033[0m'

# ============================================================================
# MAIN
# ============================================================================
def main():
    try:
        arduino = serial.Serial(ARDUINO_PORT, BAUD_RATE, timeout=1)
        time.sleep(2)
        
        os.system('clear')
        print("=" * 70)
        print(f"{BOLD}{CYAN}  ARDUINO ACCESS LOGGER v2.0{RESET}")
        print("=" * 70)
        print(f"{GREEN}[+] Connected to {ARDUINO_PORT}{RESET}")
        print(f"{YELLOW}[*] Enter password on keypad (default: 6666#){RESET}")
        print("=" * 70)
        print()
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Access Logs"
        ws.append(["Timestamp", "Event"])
        wb.save(LOG_FILE)
        
        print(f"{GREEN}[+] Log file: {LOG_FILE}{RESET}")
        print("-" * 70)
        print()
        
        entry_count = 0
        
        while True:
            if arduino.in_waiting > 0:
                line = arduino.readline().decode('utf-8').strip()
                
                if line:
                    # =========================================================
                    # ASTERISKS - Show password input
                    # =========================================================
                    if line.startswith("Asterisks:"):
                        asterisks = line[10:]
                        print(f"{YELLOW}[*] Password: {asterisks}{RESET}")
                    
                    # =========================================================
                    # LOG MESSAGES - Save to Excel
                    # =========================================================
                    elif line.startswith("LOG:"):
                        log_message = line[4:]
                        log_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        entry_count += 1
                        
                        if "Granted" in log_message:
                            print(f"{GREEN}[+] [{log_time}] {log_message}{RESET}")
                        elif "Wrong" in log_message:
                            print(f"{RED}[-] [{log_time}] {log_message}{RESET}")
                        elif "Locked" in log_message:
                            print(f"{RED}[!] [{log_time}] {log_message}{RESET}")
                        elif "Reset" in log_message:
                            print(f"{CYAN}[i] [{log_time}] {log_message}{RESET}")
                        else:
                            print(f"{YELLOW}[*] [{log_time}] {log_message}{RESET}")
                        
                        ws.append([log_time, log_message])
                        wb.save(LOG_FILE)
                        
                        if entry_count % 5 == 0:
                            print(f"{CYAN}[i] Saved ({entry_count} entries){RESET}")
                    
                    # =========================================================
                    # STATUS MESSAGES
                    # =========================================================
                    elif "ACCESS GRANTED" in line:
                        print(f"{GREEN}[+] {line}{RESET}")
                    
                    elif "WRONG PASSWORD" in line:
                        print(f"{RED}[-] {line}{RESET}")
                    
                    elif "UNLOCKING" in line:
                        print(f"{GREEN}[+] {line}{RESET}")
                    
                    elif "LOCKING" in line:
                        print(f"{RED}[!] {line}{RESET}")
                    
                    elif "SYSTEM LOCKED" in line:
                        print(f"{RED}[!] {line}{RESET}")
                    
                    elif "OBJECT DETECTED" in line:
                        print(f"{YELLOW}[*] IR Sensor - OBJECT DETECTED!{RESET}")
                    
                    elif "AUTO-RESET" in line:
                        print(f"{CYAN}[i] {line}{RESET}")
                    
                    elif "STATUS:" in line:
                        print(f"{BLUE}[i] {line[7:]}{RESET}")
                    
                    # =========================================================
                    # SUBMIT
                    # =========================================================
                    elif line.startswith("SUBMIT:"):
                        print(f"{CYAN}[i] {line[7:]}{RESET}")
                    
                    # =========================================================
                    # IGNORE EVERYTHING ELSE
                    # =========================================================
                    # All other messages are ignored
        
    except serial.SerialException as e:
        print(f"{RED}[ERROR] Could not open {ARDUINO_PORT}{RESET}")
        print(f"        Try: sudo chmod 666 {ARDUINO_PORT}")
        sys.exit(1)
        
    except KeyboardInterrupt:
        print("\n" + "=" * 70)
        print(f"{YELLOW}[!] Stopped by user{RESET}")
        
    finally:
        if 'arduino' in locals() and arduino.is_open:
            arduino.close()
            print(f"{CYAN}[i] Serial connection closed{RESET}")
        
        if os.path.exists(LOG_FILE):
            print(f"{GREEN}[+] Log file: {LOG_FILE}{RESET}")
            try:
                from openpyxl import load_workbook
                wb = load_workbook(LOG_FILE)
                ws = wb.active
                total = ws.max_row - 1
                print(f"{BLUE}[i] Total entries: {total}{RESET}")
            except:
                pass
        
        print("=" * 70)

if __name__ == "__main__":
    main()
